
from openpyxl import load_workbook
import sys
import os
from dotenv import load_dotenv
from smsKioskClass import smsKiosk
load_dotenv()

def checkWorkbook():
    listName = input("Name of the file: ")
    while not os.path.isfile(listName):
        print("File not found.")
        if input("Try again? (Y/y or otherwise): ") in ["y", "Y", "yes", "Yes"]:
            listName = input("Name of the file: ")
        else:
            return

    try:
        workbookObj = load_workbook(listName)
    except Exception as error:
        print(str(error))
        raise error

    mainSheet = workbookObj.active
    count = 0
    for row in mainSheet.iter_rows(min_row=2, max_col=5, max_row=mainSheet.max_row, values_only=True):
        try:
            firstName = row[1].lower()
            capFirstName = firstName[0].upper() + firstName[1:]
            pNum = str(row[4])
            pNum = getNumberAuto(pNum)
            print(str(count) + ": " + capFirstName + " " + pNum)
            count += 1
        except Exception as error:
            break
    print(str(count))

#*******************************************************************************
#Function: main
#Purpose: Driver
#*******************************************************************************

def main():

    startupPass = True
    RINGCENTRAL_EXTENSION = '101'

    if sys.argv[1] == "sandbox":
        if sys.argv[2] == "t":
            RINGCENTRAL_CLIENTID = os.getenv("TS_CID")
            RINGCENTRAL_CLIENTSECRET = os.getenv("TS_CS")
            RINGCENTRAL_USERNAME = os.getenv("TS_PNUM")
            RINGCENTRAL_PASSWORD = os.getenv("TS_PASS")
            RINGCENTRAL_SERVER = os.getenv("SAND_ADDRESS")
        elif sys.argv[2] == "hv":
            RINGCENTRAL_CLIENTID = os.getenv("HVS_CID")
            RINGCENTRAL_CLIENTSECRET = os.getenv("HVS_CS")
            RINGCENTRAL_USERNAME = os.getenv("HVS_PNUM")
            RINGCENTRAL_PASSWORD = os.getenv("HVS_PASS")
            RINGCENTRAL_SERVER = os.getenv("SAND_ADDRESS")
        elif sys.argv[2] == "h":
            RINGCENTRAL_CLIENTID = os.getenv("HS_CID")
            RINGCENTRAL_CLIENTSECRET = os.getenv("HS_CS")
            RINGCENTRAL_USERNAME = os.getenv("HS_PNUM")
            RINGCENTRAL_PASSWORD = os.getenv("HS_PASS")
            RINGCENTRAL_SERVER = os.getenv("SAND_ADDRESS")
        elif sys.argv[2] == "evc":
            RINGCENTRAL_CLIENTID = os.getenv("EVCS_CID")
            RINGCENTRAL_CLIENTSECRET = os.getenv("EVCS_CS")
            RINGCENTRAL_USERNAME = os.getenv("EVCS_PNUM")
            RINGCENTRAL_PASSWORD = os.getenv("EVCS_PASS")
            RINGCENTRAL_SERVER = os.getenv("SAND_ADDRESS")
        else:
            print("No proper identity named")
            startupPass = False
    elif sys.argv[1] == "t":
        RINGCENTRAL_CLIENTID = os.getenv("T_CID")
        RINGCENTRAL_CLIENTSECRET = os.getenv("T_CS")
        RINGCENTRAL_USERNAME = os.getenv("T_PNUM")
        RINGCENTRAL_PASSWORD = os.getenv("T_PASS")
        RINGCENTRAL_SERVER = os.getenv("PROD_ADDRESS")
    elif sys.argv[1] == "hv":
        RINGCENTRAL_CLIENTID = os.getenv("HV_CID")
        RINGCENTRAL_CLIENTSECRET = os.getenv("HV_CS")
        RINGCENTRAL_USERNAME = os.getenv("HV_PNUM")
        RINGCENTRAL_PASSWORD = os.getenv("HV_PASS")
        RINGCENTRAL_SERVER = os.getenv("PROD_ADDRESS")
    elif sys.argv[1] == "h":
        RINGCENTRAL_CLIENTID = os.getenv("H_CID")
        RINGCENTRAL_CLIENTSECRET = os.getenv("H_CS")
        RINGCENTRAL_USERNAME = os.getenv("H_PNUM")
        RINGCENTRAL_PASSWORD = os.getenv("H_PASS")
        RINGCENTRAL_SERVER = os.getenv("PROD_ADDRESS")
    elif sys.argv[1] == "evc":
        RINGCENTRAL_CLIENTID = os.getenv("EVC_CID")
        RINGCENTRAL_CLIENTSECRET = os.getenv("EVC_CS")
        RINGCENTRAL_USERNAME = os.getenv("EVC_PNUM")
        RINGCENTRAL_PASSWORD = os.getenv("EVC_PASS")
        RINGCENTRAL_SERVER = os.getenv("PROD_ADDRESS")
    else:
        print("No proper identity named")
        startupPass = False

    frontDesk = smsKiosk(RINGCENTRAL_CLIENTID, RINGCENTRAL_CLIENTSECRET, RINGCENTRAL_USERNAME, RINGCENTRAL_PASSWORD, RINGCENTRAL_SERVER, RINGCENTRAL_EXTENSION)

    while True:
        print("\n1. Send a single message\n2. Send a group message\n3. Get messages\n4. Check workbook\n5. Exit process\n")
        uChoice = input("Pick an option: ")
        if int(uChoice) == 1:
            frontDesk.sendSingleMessage()
        elif int(uChoice) == 2:
            try:
                frontDesk.sendBulkSMS()
            except Exception as error:
                print(str(error))
        elif int(uChoice) == 3:
            print(sPhone.getMessages('Outbound', '2020-05-25T22:00:00.000Z'))
        elif int(uChoice) == 4:
            checkWorkbook()
        elif int(uChoice) == 5:
            print("Goodbye!")
            break
        else:
            print("No valid option chosen\n")
        frontDesk.oPhone.resetThrottles()

    del frontDesk

if __name__ == "__main__":
    main()
