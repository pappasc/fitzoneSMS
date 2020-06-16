import phonenumbers
import datetime
import math
import time
import os
import multiprocessing
from openpyxl import load_workbook
from studioPhoneClass import studioPhone
from smsMessageClass import smsMessage

def getNumber():
    while True:
        userNum = input("What is the phone number (start with area code)?: ")
        parseNum = phonenumbers.parse(str(userNum), "US")
        if phonenumbers.is_valid_number(parseNum):
            return phonenumbers.format_number(parseNum, phonenumbers.PhoneNumberFormat.E164)
        else:
            print("Invalid number. Try again")

def getNumberAuto(pfNum):
    index = len(pfNum)
    for i in range(len(pfNum)):
        if pfNum[i] == ".":
            index = i
    pNum = pfNum[:index]

    try:
        parseNum = phonenumbers.parse(str(pNum), "US")
        if phonenumbers.is_valid_number(parseNum):
            return phonenumbers.format_number(parseNum, phonenumbers.PhoneNumberFormat.E164)
        else:
            raise Exception("Number not valid")
    except Exception as error:
        print("Error in getNumberAuto: " + str(error))
        raise error

def worker_task(ident, oPhone, fiName, checkList):
    time.sleep(10)
    print("\n****** Worker " + str(ident) + " is starting to recheck " + str(len(checkList)) + " messages ******\n")

    for teMess in checkList:
        readOutMess = "Worker " + str(ident) + ": Message " + str(teMess.getMessId()) + " to number " + teMess.getPnum() + " to: " + teMess.getRecipient()
        while True:
            try:
                oPhone.checkMessageStatus(teMess)
            except Exception as error:
                print("Worker " + str(ident) + ": Error checking message status: " + str(error))
                #not raising error because I want it to keep going
            else:
                sentResult = teMess.getResult()
                break
        if sentResult == 'Sent' or sentResult == 'Delivered' or sentResult == 'Received':
            outcomeMess = " was sent"
        elif sentResult == 'SendingFailed' or sentResult == 'DeliveryFailed':
            outcomeMess = " failed."
            if teMess.getSendAttempt() < 10:
                while True:
                    try:
                        oPhone.sendSMS(teMess)
                    except Exception as error:
                        print("Worker " + str(ident) + ": Error sending message: " + str(error))
                        #not raising error because I want it to keep going
                    else:
                        outcomeMess = " Resent"
                        checkList.append(teMess)
                        break
            else:
                outcomeMess = " absolutely failed"
                f = open(fiName, "a")
                f.write("Failed: " + teMess.getPnum() + " " + teMess.getRecipient() + "\n")
                f.close()
        elif sentResult == 'Queued':
            if teMess.getReQueue() > 10:
                outcomeMess = " requeued a lot"
                f = open(fiName, "a")
                f.write("Queued: " + teMess.getPnum() + " " + teMess.getRecipient() + "\n")
                f.close()
            else:
                outcomeMess = " requeuing"
                teMess.increaseReQueue()
                checkList.append(teMess)

        print(readOutMess + outcomeMess)

    print("Worker " + str(ident) + " is finished")

#*******************************************************************************
#Class: smsKiosk
#Purpose: Class  handling bulk message sending
#*******************************************************************************

class smsKiosk:
    def __init__(self, rcId, rcCS, rcUn, rcPass, rcServ, rcExt):
        self.oPhone = studioPhone(rcId, rcCS, rcUn, rcPass, rcServ, rcExt)
        numWorkers = 0
        while numWorkers < 1 or numWorkers > 10:
            try:
                numWorkers = int(input("Number of worker processes (1 to half as many clients): "))
            except Exception as error:
                print(str(error))
        self.numWorkers = numWorkers
        self.oPhone.checkInterval = math.ceil((6/5)*numWorkers)

    def sendSingleMessage(self):
        try:
            cuPnum = getNumber()
            uMessage = input("Type the message you want to send: ")
            self.oPhone.sendSMS(cuPnum, uMessage, 0)
        except Exception as error:
            print("Exception in singleMessOption: " + str(error))

    def sendBulkSMS(self):

        x = datetime.datetime.now()
        fName = x.strftime("%y%b%d_%H_%M_%S_failedMessageNums.out")

        listName = input("Name of the member list file: ")
        while not os.path.isfile(listName):
            print("File not found.")
            if input("Try again? (Y/y or otherwise): ") in ["y", "Y", "yes", "Yes"]:
                listName = input("Name of the member list file: ")
            else:
                return

        try:
            workbookObj = load_workbook(listName)
        except Exception as error:
            print(str(error))
            return

        mainSheet = workbookObj.active

        while True:
            messageFileName = input("Name of the message file: ")
            while not os.path.isfile(messageFileName):
                print("File not found.")
                if input("Try again? (Y/y or otherwise): ") in ["y", "Y", "yes", "Yes"]:
                    messageFileName = input("Name of the message file: ")
                else:
                    return
            try:
                messageFile = open(messageFileName, "r")
                messageBody = messageFile.readline()
            except Exception as error:
                print(str(error))
                raise

            print("Message in file:\n")
            print(messageBody)
            if input("Is this the correct message? (Y/y or otherwise): ") in ["y", "Y", "yes", "Yes"]:
                break

        recheckMessages = []
        for i in range(self.numWorkers):
            recheckMessages.append([])
        successCount = 0
        count = 1

        #Iterate through the spreadsheet
        for row in mainSheet.iter_rows(min_row=2, max_col=5, max_row=mainSheet.max_row, values_only=True):
            if successCount > 5:
                self.oPhone.decreaseSendThrottle()
            try:
                #Get the information from the spreadsheet, parse it, construct message
                firstName = row[1].lower()
                capFirstName = firstName[0].upper() + firstName[1:]
                lastName = row[0].lower()
                capLastName = lastName[0].upper() + lastName[1:]
                fullName = capFirstName + " " + capLastName
                pNum = getNumberAuto(str(row[4]))
            except Exception as error:
                print(str(error))
                break

            tMessage = messageBody
            textMessage = smsMessage(pNum, tMessage, fullName)

            #Attempt sending the message
            try:
                self.oPhone.sendSMS(textMessage)
            except:
                print("Error sending message in bulk send to " + fullName + " at " + pNum)
            else:
                print("Message attempt sent to " + fullName + " at " + pNum)
                recheckMessages[count % self.numWorkers].append(textMessage)
                count = count + 1
                successCount += 1
                time.sleep(self.oPhone.sendInterval)

        print("Bulk text finished!")

        def start(task, *args):
            process = multiprocessing.Process(target=task, args=args)
            process.daemon = True
            process.start()
        for i in range(self.numWorkers):
            start(worker_task, i, self.oPhone, fName, recheckMessages[i])
